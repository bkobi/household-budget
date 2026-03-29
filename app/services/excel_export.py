import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import get_db
from app.models import CATEGORIES

_thin   = Side(style="thin", color="FFCCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _h(ws, row, col, val, bg="FF217346", fg="FFFFFFFF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color=fg, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _border
    return c

def _d(ws, row, col, val, fmt=None, bg=None, align="right"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _border
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill("solid", fgColor=bg)
    return c

MONTHS_HE = ["","ינואר","פברואר","מרץ","אפריל","מאי","יוני",
             "יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

def export_excel(month: int, year: int, summary: dict) -> str:
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "סיכום חודשי"
    ws1.sheet_view.rightToLeft = True

    ws1["A1"] = f"דוח משק בית – {MONTHS_HE[month]} {year}"
    ws1["A1"].font = Font(bold=True, size=14, name="Arial")
    ws1.merge_cells("A1:E1")
    ws1["A1"].alignment = Alignment(horizontal="center")

    for col, (label, val, fmt) in enumerate([
        ("הכנסה",  summary["income"],      '#,##0.00 ₪'),
        ("הוצאות", summary["total_spent"], '#,##0.00 ₪'),
        ("יתרה",   "=B4-C4",              '#,##0.00 ₪'),
    ], 1):
        _h(ws1, 3, col, label)
        c = ws1.cell(row=4, column=col,
                     value=summary["income"] if col==1 else summary["total_spent"] if col==2 else "=B4-C4")
        c.number_format = fmt
        c.font   = Font(name="Arial")
        c.border = _border
        c.alignment = Alignment(horizontal="center")

    hdrs = ["קטגוריה", "תקציב (₪)", "בפועל (₪)", "הפרש (₪)", "ניצול %"]
    for col, h in enumerate(hdrs, 1): _h(ws1, 6, col, h)

    for i, cat in enumerate(summary["categories"]):
        r  = 7 + i
        bg = "FFF2F2F2" if i % 2 == 0 else "FFFFFFFF"
        over = cat["budget"] > 0 and cat["spent"] > cat["budget"]
        rb   = "FFFDE8E8" if over else bg
        _d(ws1, r, 1, cat["name"],   bg=rb, align="right")
        _d(ws1, r, 2, cat["budget"], fmt='#,##0.00 ₪', bg=rb)
        _d(ws1, r, 3, cat["spent"],  fmt='#,##0.00 ₪', bg=rb)
        diff = ws1.cell(row=r, column=4, value=f"=B{r}-C{r}")
        diff.number_format = '#,##0.00 ₪'
        diff.font   = Font(name="Arial", color="FFE24B4A" if over else "FF217346")
        diff.border = _border; diff.fill = PatternFill("solid", fgColor=rb)
        diff.alignment = Alignment(horizontal="right")
        pct = ws1.cell(row=r, column=5, value=f"=IF(B{r}>0,C{r}/B{r},0)")
        pct.number_format = "0.0%"; pct.border = _border
        pct.fill = PatternFill("solid", fgColor=rb)
        pct.alignment = Alignment(horizontal="center")

    tr = 7 + len(summary["categories"])
    _h(ws1, tr, 1, 'סה"כ', bg="FF333333")
    _h(ws1, tr, 2, f"=SUM(B7:B{tr-1})", bg="FF333333")
    _h(ws1, tr, 3, f"=SUM(C7:C{tr-1})", bg="FF333333")
    _h(ws1, tr, 4, f"=B{tr}-C{tr}",     bg="FF333333")
    _h(ws1, tr, 5, "",                  bg="FF333333")
    for col, w in zip(range(1,6), [22,16,16,16,12]):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A7"

    # Sheet 2 – transactions
    ws2 = wb.create_sheet("פירוט הוצאות")
    ws2.sheet_view.rightToLeft = True
    conn = get_db()
    txns = conn.execute(
        "SELECT * FROM transactions WHERE strftime('%m',date)=? AND strftime('%Y',date)=? ORDER BY date",
        (f"{month:02d}", str(year))
    ).fetchall()
    conn.close()

    for col, h in enumerate(["תאריך","תיאור","קטגוריה","סכום (₪)"], 1): _h(ws2, 1, col, h)
    for i, t in enumerate(txns):
        r  = 2 + i
        bg = "FFF2F2F2" if i % 2 == 0 else "FFFFFFFF"
        cat_name = next((c["name"] for c in CATEGORIES if c["id"] == t["category"]), t["category"])
        _d(ws2, r, 1, t["date"],        bg=bg, align="center")
        _d(ws2, r, 2, t["description"], bg=bg, align="right")
        _d(ws2, r, 3, cat_name,         bg=bg, align="right")
        _d(ws2, r, 4, t["amount"],      fmt='#,##0.00 ₪', bg=bg)
    sr = 2 + len(txns)
    _h(ws2, sr, 3, 'סה"כ',           bg="FF333333")
    _h(ws2, sr, 4, f"=SUM(D2:D{sr-1})", bg="FF333333")
    for col, w in zip(range(1,5), [14,32,18,16]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    path = os.path.join(tempfile.gettempdir(), f"budget_{year}_{month:02d}.xlsx")
    wb.save(path)
    return path
